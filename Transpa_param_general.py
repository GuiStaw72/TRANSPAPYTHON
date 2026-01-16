
#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Génère param_general.xlsx (onglet 'input') à partir d'un répertoire (SharePoint synchro locale).

Règles :
- Choix de l'onglet (Excel) : onglet ayant le + grand nombre de LIGNES où (A non vide) OU (B non vide),
  en cas d'égalité, on prend le plus à gauche.
- ISIN Excel : lire exclusivement A2 DE L’ONGLET SÉLECTIONNÉ → ISIN si valide, sinon 'err'+A2, sinon ''.
- ISIN CSV : lire exclusivement A2 (ligne 2, colonne 1) → ISIN si valide, sinon 'err'+A2, sinon ''.

Colonnes du fichier de sortie :
- fichier | Type | séparateur (décimal pour CSV) | Onglet | ISIN
"""

import argparse
import re
from pathlib import Path
from typing import Any, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook  # .xlsx

# .xls optionnel
try:
    import xlrd
    HAS_XLRD = True
except Exception:
    xlrd = None
    HAS_XLRD = False

ISIN_REGEX = re.compile(r'^[A-Z]{2}[A-Z0-9]{10}$')
MAX_ERR_LEN = 80  # troncature des 'err...'


# ----------------------- Utils généraux -----------------------

def is_temp_or_hidden(p: Path) -> bool:
    """Ignore les fichiers temporaires/masqués (~$, .*, .tmp)."""
    name = p.name
    return name.startswith('~$') or name.startswith('.') or name.endswith('.tmp')


def normalize_isin(value: Any) -> Optional[str]:
    """ISIN valide → retourne la chaîne normalisée ; sinon None."""
    if value is None:
        return None
    if isinstance(value, (int, float)):  # un ISIN n'est pas un nombre pur
        return None
    s = str(value).strip().upper()
    s = re.sub(r'\s+', '', s)
    return s if ISIN_REGEX.match(s) else None


def format_isin_or_err(raw: Any) -> str:
    """ISIN si valide, sinon 'err'+valeur, sinon '' si None/vide."""
    if raw is None:
        return ''
    s = str(raw).strip()
    if not s:
        return ''
    valid = normalize_isin(s)
    if valid:
        return valid
    payload = s if len(s) <= MAX_ERR_LEN else (s[:MAX_ERR_LEN] + '…')
    return f"err{payload}"


def row_has_value(v: Any) -> bool:
    """Non vide = tout sauf None ou chaîne vide (après strip)."""
    if v is None:
        return False
    if isinstance(v, str):
        return v.strip() != ''
    return True  # nombres/dates/bools → non vide


# ----------------------- CSV -----------------------

def guess_csv_decimal_separator(csv_path: Path, max_lines: int = 120) -> Optional[str]:
    """Détecte le séparateur décimal ('.' ou ',') dans un CSV par échantillonnage rapide."""
    candidate_delims = [',', ';', '\t', '|']
    encodings_to_try = ['utf-8-sig', 'latin-1', 'cp1252']

    comma_num = re.compile(r"^[+-]?(\d{1,3}([ \u00A0.\']?\d{3})+|\d+),\d+$")
    dot_num   = re.compile(r"^[+-]?(\d{1,3}([ \u00A0,']?\d{3})+|\d+)\.\d+$")

    comma_dec = dot_dec = 0
    for enc in encodings_to_try:
        try:
            with open(csv_path, 'r', encoding=enc, errors='ignore') as f:
                lines = []
                for i, line in enumerate(f):
                    if i >= max_lines:
                        break
                    s = line.strip()
                    if s:
                        lines.append(s)
            if not lines:
                return None

            # délimiteur de champs probable (pour découper grossièrement)
            delim_scores = {d: 0 for d in candidate_delims}
            for ln in lines:
                for d in candidate_delims:
                    delim_scores[d] += ln.count(d)
            field_delim = max(delim_scores, key=delim_scores.get) if sum(delim_scores.values()) > 0 else ','

            for ln in lines:
                parts = [p.strip().strip('"').strip("'") for p in ln.split(field_delim)]
                for tok in parts:
                    if not tok:
                        continue
                    tok2 = tok.replace('%', '').replace('\u00A0', ' ').strip()
                    if comma_num.match(tok2):
                        comma_dec += 1
                    elif dot_num.match(tok2):
                        dot_dec += 1

            if comma_dec == 0 and dot_dec == 0:
                return None
            return ',' if comma_dec >= dot_dec else '.'
        except Exception:
            continue
    return None


def get_csv_a2_isin(csv_path: Path) -> str:
    """Lit EXCLUSIVEMENT A2 du CSV (ligne 2, col 1) et retourne ISIN/err/vide."""
    encodings_to_try = ['utf-8-sig', 'cp1252', 'latin-1']

    # 1) Inférence auto du séparateur
    for enc in encodings_to_try:
        try:
            df = pd.read_csv(
                csv_path,
                sep=None, engine='python', encoding=enc,
                nrows=2, header=None, dtype=str, keep_default_na=False
            )
            if df.shape[0] >= 2 and df.shape[1] >= 1:
                return format_isin_or_err(df.iloc[1, 0])  # A2
            return ''
        except Exception:
            pass

    # 2) Séparateurs explicites
    for enc in encodings_to_try:
        for delim in [',', ';', '\t', '|']:
            try:
                df = pd.read_csv(
                    csv_path,
                    sep=delim, engine='python', encoding=enc,
                    nrows=2, header=None, dtype=str, keep_default_na=False
                )
                if df.shape[0] >= 2 and df.shape[1] >= 1:
                    return format_isin_or_err(df.iloc[1, 0])  # A2
                return ''
            except Exception:
                continue
    return ''


# ----------------------- Excel .xlsx -----------------------

def count_rows_with_A_or_B_xlsx(ws, max_rows: int) -> int:
    """Compte le nb de LIGNES (jusqu'à max_rows) où A OU B est non vide."""
    cnt = 0
    for a, b in ws.iter_rows(min_row=1, max_row=max_rows, min_col=1, max_col=2, values_only=True):
        if row_has_value(a) or row_has_value(b):
            cnt += 1
    return cnt


def select_xlsx_sheet_by_ab_rows(xlsx_path: Path, max_rows: int) -> Optional[str]:
    """
    Retourne le NOM de l’onglet .xlsx qui maximise le nombre de LIGNES avec (A non vide) OU (B non vide).
    Tie-break : plus à gauche.
    """
    try:
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        best_name = None
        best_count = -1
        for ws in wb.worksheets:  # gauche -> droite
            c = count_rows_with_A_or_B_xlsx(ws, max_rows)
            if c > best_count:
                best_count = c
                best_name = ws.title
        return best_name
    except Exception:
        return None


def get_xlsx_a2_isin(xlsx_path: Path, sheet_name: str) -> str:
    """Lit EXCLUSIVEMENT A2 de l’onglet sheet_name d’un .xlsx et retourne ISIN/err/vide."""
    try:
        wb = load_workbook(filename=xlsx_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return ''
        ws = wb[sheet_name]
        try:
            a2 = ws['A2'].value
        except Exception:
            a2 = None
        return format_isin_or_err(a2)
    except Exception:
        return ''


# ----------------------- Excel .xls -----------------------

def count_rows_with_A_or_B_xls(sheet, max_rows: int) -> int:
    rmax = min(sheet.nrows, max_rows)
    cnt = 0
    for r in range(rmax):
        a = sheet.cell_value(r, 0) if sheet.ncols > 0 else None
        b = sheet.cell_value(r, 1) if sheet.ncols > 1 else None
        if row_has_value(a) or row_has_value(b):
            cnt += 1
    return cnt


def select_xls_sheet_by_ab_rows(xls_path: Path, max_rows: int) -> Optional[str]:
    """Retourne le NOM de l’onglet .xls qui maximise le nb de LIGNES avec (A ou B non vide)."""
    if not HAS_XLRD:
        return None
    try:
        wb = xlrd.open_workbook(xls_path.as_posix())
        best_name = None
        best_count = -1
        for sheet in wb.sheets():  # gauche -> droite
            c = count_rows_with_A_or_B_xls(sheet, max_rows)
            if c > best_count:
                best_count = c
                best_name = sheet.name
        return best_name
    except Exception:
        return None


def get_xls_a2_isin(xls_path: Path, sheet_name: str) -> str:
    """Lit EXCLUSIVEMENT A2 de l’onglet sheet_name d’un .xls et retourne ISIN/err/vide."""
    if not HAS_XLRD:
        return ''
    try:
        wb = xlrd.open_workbook(xls_path.as_posix())
        try:
            sh = wb.sheet_by_name(sheet_name)
        except Exception:
            return ''
        try:
            a2 = sh.cell_value(1, 0)  # A2 = (row=1, col=0)
        except Exception:
            a2 = None
        return format_isin_or_err(a2)
    except Exception:
        return ''


# ----------------------- Scan du répertoire -----------------------

def scan_directory(
    folder: Path,
    csv_onglet_mode: str = "basename",
    csv_onglet_fixed_value: str = "csv",
    scan_max_rows: int = 500,
    extensions: Tuple[str, ...] = ('.csv', '.xls', '.xlsx'),
) -> List[dict]:
    """
    Construit les lignes pour le DataFrame final.
    csv_onglet_mode:
      - 'basename' -> onglet = nom du fichier CSV sans extension
      - 'fixed'    -> onglet = csv_onglet_fixed_value
      - 'empty'    -> onglet = ''
    """
    rows: List[dict] = []
    for p in sorted(folder.rglob('*')):
        if not p.is_file() or is_temp_or_hidden(p):
            continue

        ext = p.suffix.lower()
        if ext not in extensions:
            continue

        type_val = ext.lstrip('.')
        sep_val = ''
        sheet_val = ''
        isin_val = ''

        if ext == '.csv':
            # 1) séparateur décimal
            sep = guess_csv_decimal_separator(p)
            sep_val = sep if sep in (',', '.') else ''

            # 2) "Onglet" = nom du fichier sans extension (ou autre mode)
            if csv_onglet_mode == 'basename':
                sheet_val = p.stem
            elif csv_onglet_mode == 'fixed':
                sheet_val = csv_onglet_fixed_value
            else:
                sheet_val = ''

            # 3) ISIN = A2 du CSV
            isin_val = get_csv_a2_isin(p)

        elif ext == '.xlsx':
            # 1) sélectionner l’onglet via la règle A/B
            sheet_name = select_xlsx_sheet_by_ab_rows(p, scan_max_rows) or ''
            sheet_val = sheet_name

            # 2) lire A2 de CET onglet (et seulement A2)
            isin_val = get_xlsx_a2_isin(p, sheet_name) if sheet_name else ''

        elif ext == '.xls':
            sheet_name = select_xls_sheet_by_ab_rows(p, scan_max_rows) or ''
            sheet_val = sheet_name
            isin_val = get_xls_a2_isin(p, sheet_name) if sheet_name else ''

        rows.append({
            'fichier': p.name,
            'Type': type_val,
            'séparateur': sep_val,
            'Onglet': sheet_val,
            'ISIN': isin_val,
        })

    return rows


# ----------------------- Main -----------------------

def main():
    parser = argparse.ArgumentParser(
        description="Génère param_general.xlsx (onglet 'input') listant CSV/XLS/XLSX d'un répertoire SharePoint (synchro locale)."
    )
    parser.add_argument("repertoire", help="Chemin du répertoire à analyser")
    parser.add_argument("--output", default="param_general.xlsx", help="Fichier Excel de sortie")
    parser.add_argument("--sheet", default="input", help="Nom de l'onglet de sortie (défaut : input)")

    # 'Onglet' pour CSV
    parser.add_argument("--csv-onglet-mode", choices=["basename", "fixed", "empty"], default="basename",
                        help="Valeur pour la colonne 'Onglet' des CSV (défaut : basename)")
    parser.add_argument("--csv-onglet-fixed-value", default="csv",
                        help="Valeur utilisée si --csv-onglet-mode=fixed")

    # Règle Excel : nb max de lignes A/B à compter pour la sélection d’onglet
    parser.add_argument("--scan-max-rows", type=int, default=500,
                        help="Nb max de lignes comptées par feuille (défaut : 500)")

    args = parser.parse_args()

    folder = Path(args.repertoire).expanduser().resolve()
    if not folder.exists() or not folder.is_dir():
        raise SystemExit(f"Le répertoire n'existe pas ou n'est pas un dossier : {folder}")

    rows = scan_directory(
        folder=folder,
        csv_onglet_mode=args.csv_onglet_mode,
        csv_onglet_fixed_value=args.csv_onglet_fixed_value,
        scan_max_rows=args.scan_max_rows,
    )

    df = pd.DataFrame(rows, columns=['fichier', 'Type', 'séparateur', 'Onglet', 'ISIN'])

    with pd.ExcelWriter(args.output, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=args.sheet, index=False)

    if not HAS_XLRD and any(r['Type'] == 'xls' for r in rows):
        print("⚠️ xlrd n'est pas installé : les .xls sont listés mais sans lecture. Installez-le : pip install xlrd")

    print(f"✅ Fichier généré : {Path(args.output).resolve()} (onglet : {args.sheet})")
    print(f"Total de lignes : {len(df)}")


if __name__ == "__main__":
    main()





